import os
import sys
import time
import click
import polib
import openpyxl
from typing import List
from pathlib import Path
from enum import Enum, unique

from openpyxl.styles import Font, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell


class ColumnHeaders:
    message_context = "Message context"
    message_id = "Message id"
    comment_source = "Source comment"
    comment_translator = "Translator comment"
    comment_references = "References"


class PortableObjectFile:
    """ Represents a po file """

    def __init__(self, file_path, locale=None):
        self.file_path = file_path
        self.po_file = None
        self.locale = locale

        # Convert
        if not os.path.exists(self.file_path) and ":" in self.file_path:
            # The user passed a <locale>:<path> value
            self.locale, self.file_path = self.file_path.split(":", 1)
            self.file_path = Path(self.file_path).resolve()
            self.po_file = polib.pofile(self.file_path, encoding="utf-8-sig")
        else:
            self.file_path = Path(self.file_path).resolve()
            self.po_file = polib.pofile(self.file_path, encoding="utf-8-sig")

            # Fallback to metadata
            if not self.locale:
                self.locale = self.po_file.metadata.get("Language")

            # Fallback to filename without extension
            if not self.locale:
                self.locale = self.file_path.stem

    def has_any_message_context(self):
        return any(m.msgctxt for m in self.po_file)


class PortableObjectFileToXLSX:
    """
    Convert .PO files to an XLSX file.

    po-to-xls tries to guess the locale for PO files by looking at the
    "Language" key in the PO metadata, falling back to the filename. You
    can also specify the locale manually by adding prefixing the filename
    with "<locale>:". For example: "nl:locales/nl/mydomain.po".
    """

    def __init__(
        self,
        po_files: List[PortableObjectFile],
        comments_type: str,
        output_file_path: Path,
        width_message_context: int = 25,
        width_message_id: int = 100,
        width_message_locale: int = 100,
        width_comments: int = 50,
        wrap_message_id: bool = True,
        wrap_comments: bool = False,
        wrap_message_locale: bool = True,
        always_write_message_context: bool = False,
        lock_sheet: bool = False,
        font_regular_name: str = "Verdana",
        font_regular_size: int = 11,
    ):
        """
        message_context = namespace, is optional
        message_id = source string to translate
        """

        self.po_files = po_files
        self.output_file_path = output_file_path
        self.comments_type = comments_type

        # Widths should be in range [0, 200]
        self.width_message_context = width_message_context
        self.width_message_id = width_message_id
        self.width_message_locale = width_message_locale
        self.width_comments = width_comments

        # Wrap options
        self.wrap_message_id = wrap_message_id
        self.wrap_comments = wrap_comments
        self.wrap_message_locale = wrap_message_locale

        # Should we lock some cells for protection
        self.lock_sheet = lock_sheet
        self.unlock_message_locale = self.lock_sheet

        self.always_write_message_context = always_write_message_context

        self.has_message_context = False
        self.has_comment_references = False
        self.has_comment_source = False
        self.has_comment_translator = False

        # Has message context/namespace/group name
        if self.always_write_message_context:
            self.has_message_context = True
        else:
            for po_file in self.po_files:
                if self.has_message_context:
                    break
                self.has_message_context = self.has_message_context or po_file.has_any_message_context()

        # Fonts
        self.font_regular_name = font_regular_name
        self.font_regular_size = font_regular_size
        self.font_regular = Font(name=self.font_regular_name, size=self.font_regular_size)
        self.font_regular_bold = Font(name=self.font_regular_name, size=self.font_regular_size, bold=True)
        self.font_fuzzy = Font(italic=True, bold=True)

        # AlignmentAlignment
        self.alignment_wrap_text = Alignment(wrap_text=True)
        self.alignment_shrink_to_fit = Alignment(shrink_to_fit=True)

        # NOTE: using optimized mode
        self.work_book = openpyxl.Workbook(write_only=True)
        self.work_sheet = self.work_book.create_sheet(title="Translations")

        self.column_names = self.get_column_names()

        # NOTE: if we are not using optimized mode we should move this
        self.apply_style()

        self.write_columns_header()
        self.write_body()
        self.save()

    def get_column_names(self):
        columns = []

        if self.has_message_context:
            columns.append(ColumnHeaders.message_context)

        columns.append(ColumnHeaders.message_id)

        # Headers
        if "reference" in self.comments_type or "all" in self.comments_type:
            self.has_comment_references = True
            columns.append(ColumnHeaders.comment_references)
        if "extracted" in self.comments_type or "all" in self.comments_type:
            self.has_comment_source = True
            columns.append(ColumnHeaders.comment_source)
        if "translator" in self.comments_type or "all" in self.comments_type:
            self.has_comment_translator = True
            columns.append(ColumnHeaders.comment_translator)

        # The languages headers
        for f in self.po_files:
            columns.append(f.locale)

        return columns

    # NOTE: excel uses 1 base indexing
    def get_column_index_message_context(self) -> int:
        return self.column_names.index(ColumnHeaders.message_context) + 1

    def get_column_index_message_id(self) -> int:
        return self.column_names.index(ColumnHeaders.message_id) + 1

    def get_columns_indices_comments(self) -> List[int]:
        indices = []

        # index is 1 based
        if self.has_comment_references:
            indices.append(self.column_names.index(ColumnHeaders.comment_references) + 1)
        if self.has_comment_source:
            indices.append(self.column_names.index(ColumnHeaders.comment_source) + 1)
        if self.has_comment_translator:
            indices.append(self.column_names.index(ColumnHeaders.comment_translator) + 1)

        return indices

    def get_column_indices_locales(self) -> List[int]:
        indices = []

        for f in self.po_files:
            try:
                indices.append(self.column_names.index(f.locale) + 1)
            except ValueError:
                # Locale does not exist
                pass

        return indices

    def get_column_letter_message_context(self) -> str:
        return get_column_letter(self.get_column_index_message_context())

    def get_column_letter_message_id(self):
        return get_column_letter(self.get_column_index_message_id())

    def get_column_message_context(self):
        return self.work_sheet.column_dimensions[self.get_column_letter_message_context()]

    def get_column_message_id(self):
        return self.work_sheet.column_dimensions[self.get_column_letter_message_id()]

    def apply_style(self):
        # NOTE: Because we are using optimized mode we must set these before writing anything
        # https://openpyxl.readthedocs.io/en/stable/optimized.html
        # Reference: https://automatetheboringstuff.com/chapter12/

        # Lock
        if self.lock_sheet:
            self.work_sheet.protection.sheet = True

        #
        # Set sizes
        #

        # Message context and id
        column_message_context = self.get_column_message_context()
        column_message_id = self.get_column_message_id()

        column_message_context.width = self.width_message_context
        column_message_id.width = self.width_message_id

        # Comments
        for i in self.get_columns_indices_comments():
            self.work_sheet.column_dimensions[get_column_letter(i)].width = self.width_comments

        # Locales, set the width the same as the message id, as that is the source string
        for i in self.get_column_indices_locales():
            self.work_sheet.column_dimensions[get_column_letter(i)].width = self.width_message_locale

        # Freeze the first row
        self.work_sheet.freeze_panes = "A2"

        # Freeze the first 2 columns
        self.work_sheet.freeze_panes = "C2"

        # Set fonts extend to the right + 5
        for i in range(len(self.column_names) + 5):
            # index is 1 based
            self.work_sheet.column_dimensions[get_column_letter(i + 1)].font = self.font_regular

    def get_cell(self, value, wrap=False, shrink_to_fit=False, bold=False, unlock=False) -> WriteOnlyCell:
        cell = WriteOnlyCell(self.work_sheet, value=value)

        if bold:
            cell.font = self.font_regular_bold
        else:
            cell.font = self.font_regular

        if wrap:
            cell.alignment = self.alignment_wrap_text
        elif shrink_to_fit:
            cell.alignment = self.alignment_shrink_to_fit

        if unlock:
            cell.protection = Protection(locked=False)

        return cell

    def write_columns_header(self):
        row = []
        for name in self.column_names:
            row.append(self.get_cell(name, bold=True))

        self.work_sheet.append(row)

    def write_body(self):
        # Collect the messages
        messages = []
        seen = set()
        for f in self.po_files:
            for msg in f.po_file:
                # Has message
                if not msg.msgid or msg.obsolete:
                    continue

                if (msg.msgid, msg.msgctxt) not in seen:
                    messages.append((msg.msgid, msg.msgctxt))
                    seen.add((msg.msgid, msg.msgctxt))

        # used to write the first columns
        reference_po_file = self.po_files[0].po_file

        # The rest of the rows
        for msgid, msgctxt in messages:
            row = []

            # Message context
            if self.has_message_context:
                row.append(self.get_cell(msgctxt))

            # Message id
            row.append(self.get_cell(msgid, wrap=self.wrap_message_id))

            msg = reference_po_file.find(msgid, msgctxt=msgctxt)

            # Metadata comment columns
            if self.has_comment_references:
                data = []
                if msg is not None:
                    for (entry, lineno) in msg.comment_references:
                        if lineno:
                            data.append("%s:%s" % (entry, lineno))
                        else:
                            data.append(entry)

                if data:
                    row.append(
                        self.get_cell(", ".join(data), wrap=self.wrap_comments, shrink_to_fit=not self.wrap_comments)
                    )
                else:
                    row.append(self.get_cell(None, wrap=self.wrap_comments, shrink_to_fit=not self.wrap_comments))

            if self.has_comment_source:
                data = None
                if msg is not None:
                    data = msg.comment
                row.append(self.get_cell(data, wrap=self.wrap_comments, shrink_to_fit=not self.wrap_comments))

            if self.has_comment_translator:
                data = None
                if msg is not None:
                    data = msg.tcomment
                row.append(self.get_cell(data, wrap=self.wrap_comments, shrink_to_fit=not self.wrap_comments))

            # Write the language rows, aka strings to translate
            for f in self.po_files:
                po_file = f.po_file
                msg = po_file.find(msgid, msgctxt=msgctxt)
                if msg is None:
                    row.append(self.get_cell(None, wrap=self.wrap_message_locale, unlock=self.unlock_message_locale))
                elif "fuzzy" in msg.flags:
                    # Weird case
                    cell = WriteOnlyCell(self.work_sheet, value=msg.msgstr)
                    cell.font = self.font_fuzzy
                    row.append(cell, unlock=self.unlock_message_locale)
                else:
                    # Normal case
                    row.append(
                        self.get_cell(msg.msgstr, wrap=self.wrap_message_locale, unlock=self.unlock_message_locale)
                    )

            self.work_sheet.append(row)

    def save(self):
        self.work_book.save(str(self.output_file_path))
